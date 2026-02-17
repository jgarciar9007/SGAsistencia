from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django import forms
from django.utils.html import format_html
import json

from .models import Empleado, Candidato, Documento, BajaAutorizada

class ImportCandidatosForm(forms.Form):
    json_file = forms.FileField(label="Archivo JSON")


class DocumentoInline(admin.TabularInline):
    model = Documento
    extra = 1

class BajaAutorizadaInline(admin.TabularInline):
    model = BajaAutorizada
    extra = 1

import openpyxl
from datetime import datetime

class ImportEmpleadosForm(forms.Form):
    excel_file = forms.FileField(label="Archivo Excel (.xlsx)")

@admin.register(Empleado)
class EmpleadoAdmin(admin.ModelAdmin):
    # ... (rest of EmpleadoAdmin)
    list_display = (
        "numero",
        "nombre",
        "apellido",
        "departamento",
        "area",
        "tipo_vinculacion",
        "puesto",
        "activo",
    )
    list_filter = (
        "tipo_vinculacion",
        "departamento",
        "area",
        "activo",
    )
    search_fields = (
        "numero",
        "nombre",
        "apellido",
        "doc_id",
        "user_id",
        "uid",
    )
    list_editable = ("activo",)
    ordering = ("apellido", "nombre")
    
    inlines = [DocumentoInline]
    change_list_template = "admin/empleados/empleado/change_list.html"

    fieldsets = (
        ("Identificación", {
            "fields": ("numero", "doc_id", "nombre", "apellido", "foto", "tipo_vinculacion", "puesto", "salario_base")
        }),
        ("Contacto", {
            "fields": ("telefono", "email", "direccion")
        }),
        ("Organización", {
            "fields": ("departamento", "area", "seccion")
        }),
        ("Vinculación con dispositivo", {
            "fields": ("dispositivo", "user_id", "uid", "activo")
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='empleados_empleado_import_excel'),
            path('download-template/', self.admin_site.admin_view(self.download_template_view), name='empleados_empleado_download_template'),
        ]
        return my_urls + urls

    def import_excel_view(self, request):
        if request.method == "POST":
            form = ImportEmpleadosForm(request.POST, request.FILES)
            if form.is_valid():
                f = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(f)
                    ws = wb.active
                    
                    count_created = 0
                    count_updated = 0
                    
                    # Asumimos que la primera fila es encabezado
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # Mapeo basado en imagen:
                        # 0: NOMBRE COMPLETO
                        # 1: ESTADO
                        # 2: SECTOR/DEPARTAMENTO
                        # 3: NIVEL ACADEMICO (Ignorar/Nota?)
                        # 4: ESPECIALIDAD (Ignorar?)
                        # 5: PAIS (Ignorar?)
                        # 6: IDIOMAS (Ignorar?)
                        # 7: FECHA DE ALTA
                        # 8: EDAD
                        # 9: TELÉFONO
                        
                        if not row[0]: # Si no hay nombre, saltar
                            continue

                        nombre_completo = str(row[0]).strip()
                        parts = nombre_completo.split(" ", 1)
                        if len(parts) > 1:
                            nombre = parts[0]
                            apellido = parts[1]
                        else:
                            nombre = parts[0]
                            apellido = ""
                        
                        estado_raw = str(row[1] or "").upper()
                        tipo_vinc = "FUNC" # Default
                        if "CONTRAT" in estado_raw:
                            tipo_vinc = "CONT"
                        elif "PRACT" in estado_raw:
                            tipo_vinc = "PRAC"
                        elif "FUNCION" in estado_raw:
                            tipo_vinc = "FUNC"
                        
                        departamento = str(row[2] or "").strip()
                        fecha_alta_raw = row[7]
                        telefono = str(row[9] or "").strip()
                        
                        # Generar doc_id temporal si no existe o usar uno ficticio consistente
                        # Usaremos una lógica determinista simple para evitar duplicados en pruebas
                        # En producción idealmente el excel traería DNI.
                        # Aquí usaremos HASH o algo simple del nombre si no hay otra cosa.
                        # Mejor: MIGRACION-{hash(nombre_completo)}
                        doc_id = f"MIG-{abs(hash(nombre_completo))}"[:20]

                        # Numero nómina: igual
                        numero = f"N-{abs(hash(nombre_completo))}"[:10]

                        defaults = {
                            "nombre": nombre[:60],
                            "apellido": apellido[:60],
                            "departamento": departamento[:80],
                            "tipo_vinculacion": tipo_vinc,
                            "telefono": telefono[:20],
                            "numero": numero,
                            "salario_base": 0.00 # Default
                        }
                        
                        obj, created = Empleado.objects.update_or_create(
                            doc_id=doc_id,
                            defaults=defaults
                        )
                        
                        if created:
                            count_created += 1
                        else:
                            count_updated += 1

                    messages.success(request, f"Importación Excel completada: {count_created} creados, {count_updated} actualizados.")
                    return redirect("admin:empleados_empleado_changelist")

                except Exception as e:
                    messages.error(request, f"Error procesando archivo Excel: {str(e)}")
        else:
            form = ImportEmpleadosForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "form": form,
            "title": "Importar Empleados desde Excel"
        }
        return render(request, "admin/empleados/empleado/import_form.html", context)

    def download_template_view(self, request):
        from django.http import FileResponse
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empleados Template"
        
        # Headers matching the import expectations
        headers = ["NOMBRE COMPLETO", "ESTADO", "SECTOR/DEPARTAMENTO", "NOTA (OPCIONAL)", "ESPECIALIDAD", "PAIS", "IDIOMAS", "FECHA ALTA (YYYY-MM-DD)", "EDAD", "TELEFONO"]
        ws.append(headers)
        
        # Sample row
        ws.append(["Juan Perez", "Funcionario", "IT", "", "", "", "", "2024-01-01", "30", "555-0101"])
        
        # Save to memory buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return FileResponse(
            buffer, 
            as_attachment=True, 
            filename='plantilla_empleados.xlsx'
        )

@admin.register(Candidato)
class CandidatoAdmin(admin.ModelAdmin):
    list_display = ("nombre", "apellido", "doc_id", "estado", "email", "telefono")
    list_filter = ("estado",)
    search_fields = ("nombre", "apellido", "doc_id", "skills")
    inlines = [DocumentoInline]
    change_list_template = "admin/empleados/candidato/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-json/', self.admin_site.admin_view(self.import_json_view), name='empleados_candidato_import_json'),
        ]
        return my_urls + urls

    def import_json_view(self, request):
        if request.method == "POST":
            form = ImportCandidatosForm(request.POST, request.FILES)
            if form.is_valid():
                f = request.FILES["json_file"]
                try:
                    data = json.load(f)
                    # Si es una lista, iteramos
                    if isinstance(data, list):
                        count_created = 0
                        count_updated = 0
                        for item in data:
                            # Mapeo de campos. Se busca por doc_id para no duplicar.
                            # Ajusta las claves según el JSON esperado.
                            doc_id = str(item.get("doc_id", "")).strip()
                            if not doc_id:
                                continue # Skip si no hay doc_id

                            defaults = {
                                "nombre": item.get("nombre", "")[:60],
                                "apellido": item.get("apellido", "")[:60],
                                "email": item.get("email", ""),
                                "telefono": item.get("telefono", "")[:20],
                                "skills": item.get("skills", ""),
                                "titulaciones": item.get("titulaciones", ""),
                                "nota": item.get("nota", ""),
                            }

                            obj, created = Candidato.objects.update_or_create(
                                doc_id=doc_id,
                                defaults=defaults
                            )
                            if created:
                                count_created += 1
                            else:
                                count_updated += 1
                        
                        messages.success(request, f"Importación completada: {count_created} creados, {count_updated} actualizados.")
                        return redirect("admin:empleados_candidato_changelist")
                    else:
                        messages.error(request, "El JSON debe ser una lista de objetos.")
                except json.JSONDecodeError:
                    messages.error(request, "Error decodificando el archivo JSON.")
                except Exception as e:
                    messages.error(request, f"Error procesando archivo: {str(e)}")
        else:
            form = ImportCandidatosForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "form": form,
            "title": "Importar Candidatos desde JSON"
        }
        return render(request, "admin/empleados/candidato/import_form.html", context)


@admin.register(Documento)
class DocumentoAdmin(admin.ModelAdmin):
    list_display = ("tipo", "descripcion", "empleado", "candidato", "subido_en")
    list_filter = ("tipo", "subido_en")
    search_fields = ("descripcion", "empleado__nombre", "candidato__nombre")


@admin.register(BajaAutorizada)
class BajaAutorizadaAdmin(admin.ModelAdmin):
    list_display = ("empleado", "fecha_inicio", "fecha_fin", "tipo")
    list_filter = ("tipo", "fecha_inicio")
    search_fields = ("empleado__nombre", "empleado__apellido", "descripcion")
